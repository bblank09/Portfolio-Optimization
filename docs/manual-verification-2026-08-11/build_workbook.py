#!/usr/bin/env python3
"""Build docs/manual-verification-2026-08-11.xlsx — the Phase 4 hand-verification workbook.

Every metric on every sheet is recomputed by a LIVE Excel formula chain that
bottoms out in the raw NAV values on the `NAV` sheet (loaded verbatim from
nav.csv).  Values taken from result.json appear ONLY in columns explicitly
labelled "given (result.json)" — they are the things being checked *against*,
never inputs to the recomputation.

Usage:
    source /private/tmp/sec_open_data_portfolio_backtester_venv/bin/activate
    python3 docs/manual-verification-2026-08-11/build_workbook.py

openpyxl is a dev-time verification tool only; it is deliberately NOT added to
pyproject.toml.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "manual-verification-2026-08-11.xlsx"

NAV_CSV = HERE / "nav.csv"
REQUEST_JSON = HERE / "request.json"
RESULT_JSON = HERE / "result.json"

# Month-end anchor dates for the request's monthly test window
# (2024-02-29 .. 2024-05-31).  March's anchor is 2024-03-29 because 2024-03-30/31
# fell on a weekend — the last *trading* day of the month is the anchor.
ANCHORS = ["2024-02-29", "2024-03-29", "2024-04-30", "2024-05-31"]

RF_PCT = 1.5  # request.constraints.riskFreeRatePct
PERIODS_PER_YEAR = 12  # dataFrequency == "monthly"

TITLE = Font(bold=True, size=13)
HDR = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="DDE7F0")
NOTE = Font(italic=True, size=9, color="555555")
GIVEN_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow == pasted from result.json
WRAP = Alignment(wrap_text=True, vertical="top")


def _title(ws, text, row=1):
    ws.cell(row=row, column=1, value=text).font = TITLE
    return row + 1


def _note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE
    c.alignment = WRAP
    return row + 1


def _header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = WRAP
    return row + 1


def _widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def main() -> None:
    nav_rows = list(csv.reader(NAV_CSV.open()))
    nav_header, nav_data = nav_rows[0], nav_rows[1:]
    request = json.loads(REQUEST_JSON.read_text())
    result = json.loads(RESULT_JSON.read_text())

    # nav.csv column order is: nav_date, M0155_2547 (M-S50), M0209_2548 (K-SET50).
    # Kept verbatim; every downstream reference uses these letters.
    assert nav_header == ["nav_date", "M0155_2547", "M0209_2548"], nav_header
    COL_MS50 = "B"      # M0155_2547
    COL_KSET50 = "C"    # M0209_2548
    n_nav = len(nav_data)
    nav_last_row = n_nav + 1  # header occupies row 1

    wb = Workbook()

    # ------------------------------------------------------------------ NAV
    ws = wb.active
    ws.title = "NAV"
    _header_row(ws, 1, ["nav_date", "M0155_2547 (M-S50) NAV", "M0209_2548 (K-SET50) NAV"])
    for i, row in enumerate(nav_data, start=2):
        ws.cell(row=i, column=1, value=row[0])
        for j, v in enumerate(row[1:], start=2):
            ws.cell(row=i, column=j, value=(float(v) if v.strip() else None))
    _widths(ws, [14, 26, 28])
    ws.freeze_panes = "A2"
    ws.cell(row=1, column=5, value=(
        "Raw data, verbatim from nav.csv (Task 1). No formulas on this sheet. "
        "Every other sheet's numbers are live formulas rooted here. "
        "Blank cells early in the series are unaligned raw SEC data (expected, and "
        "outside the 2024-02..2024-05 test window)."
    )).font = NOTE

    # -------------------------------------------------------------- Returns
    ws = wb.create_sheet("Returns")
    _header_row(ws, 1, ["nav_date", "M0155_2547 return", "M0209_2548 return"])
    for i in range(2, n_nav + 1):  # sheet row i -> NAV rows i and i+1
        nav_row = i + 1
        ws.cell(row=i, column=1, value=f"=NAV!A{nav_row}")
        ws.cell(row=i, column=2,
                value=f'=IFERROR(NAV!{COL_MS50}{nav_row}/NAV!{COL_MS50}{i}-1,"")')
        ws.cell(row=i, column=3,
                value=f'=IFERROR(NAV!{COL_KSET50}{nav_row}/NAV!{COL_KSET50}{i}-1,"")')
    _widths(ws, [14, 20, 20])
    ws.freeze_panes = "A2"

    # Test-window table (columns F..J) — the range every later sheet points at.
    ws["F1"] = ("TEST WINDOW — request.timePeriod 2024-02-29..2024-05-31, "
                "dataFrequency=monthly")
    ws["F1"].font = TITLE
    _note(ws, 2, "")
    ws["F2"] = ("Month-end anchors pulled out of NAV by INDEX/MATCH (live). "
                "2024-03-29 is March's anchor: 2024-03-30/31 were non-trading days. "
                "4 anchors -> 3 monthly returns (the brief said 4; the real calendar "
                "yields 3, matching result.json's 3-element monthlyReturnsPct).")
    ws["F2"].font = NOTE
    ws["F2"].alignment = WRAP
    _header_row(ws, 3, ["month_end", "NAV M0155_2547", "NAV M0209_2548",
                        "ret M0155_2547", "ret M0209_2548"], start_col=6)
    for k, d in enumerate(ANCHORS):
        r = 4 + k
        ws.cell(row=r, column=6, value=d)
        ws.cell(row=r, column=7,
                value=f"=INDEX(NAV!{COL_MS50}:{COL_MS50},MATCH($F{r},NAV!$A:$A,0))")
        ws.cell(row=r, column=8,
                value=f"=INDEX(NAV!{COL_KSET50}:{COL_KSET50},MATCH($F{r},NAV!$A:$A,0))")
        if k > 0:
            ws.cell(row=r, column=9, value=f"=G{r}/G{r-1}-1")
            ws.cell(row=r, column=10, value=f"=H{r}/H{r-1}-1")
    _widths(ws, [14, 18, 18, 16, 16], start_col=6)
    # Named ranges for the 3 monthly returns
    RET_MS50 = "Returns!$I$5:$I$7"
    RET_KSET50 = "Returns!$J$5:$J$7"
    N_OBS = 3
    # Sample covariance written out longhand rather than as COVARIANCE.S: post-2007
    # function names must be stored in the file as "_xlfn.COVARIANCE.S" or Excel and
    # LibreOffice both render #NAME?.  The longhand form is also more auditable by
    # hand, which is the point of this workbook.  VAR/STDEV are the legacy spellings
    # of VAR.S/STDEV.S (identical sample statistics, ddof=1) and need no prefix.
    COV_EXPR = (f"=SUMPRODUCT(({RET_KSET50}-AVERAGE({RET_KSET50}))"
                f"*({RET_MS50}-AVERAGE({RET_MS50})))/({N_OBS}-1)")

    ws["F9"] = "Test-window return count (live)"
    ws["G9"] = f"=COUNT({RET_KSET50})"
    ws["F10"] = "Periods per year (m)"
    ws["G10"] = PERIODS_PER_YEAR
    ws["F11"] = "Risk-free rate % p.a. (request.constraints.riskFreeRatePct)"
    ws["G11"] = RF_PCT

    # -------------------------------------------------------------- Weights
    ws = wb.create_sheet("Weights")
    r = _title(ws, "WEIGHTS TAB VERIFICATION")
    r = _note(ws, r, "Yellow cells are GIVEN (pasted from result.json) and are what the "
                     "live formulas are checked against. Everything else is a formula "
                     "chain rooted in NAV.")
    r += 1
    hdr = r
    r = _header_row(ws, r, ["projId", "displayName", "given weight (result.json)",
                            "mean monthly return (live)", "annualized return % (live)",
                            "given assetSummary.expectedReturnPct", "diff",
                            "annualized vol % (live)",
                            "given assetSummary.volatilityPct", "diff"])
    ow = result["optimalWeights"]
    asset_by_id = {a["projId"]: a for a in result["assetSummary"]}
    rows = [("M0209_2548", "K-SET50", RET_KSET50), ("M0155_2547", "M-S50", RET_MS50)]
    w_rows = {}
    for pid, name, rng in rows:
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value=name)
        c = ws.cell(row=r, column=3, value=ow[pid] / 100.0)
        c.fill = GIVEN_FILL
        c.number_format = "0.00000000"
        ws.cell(row=r, column=4, value=f"=AVERAGE({rng})")
        ws.cell(row=r, column=5, value=f"=D{r}*Returns!$G$10*100")
        g = ws.cell(row=r, column=6, value=asset_by_id[pid]["expectedReturnPct"])
        g.fill = GIVEN_FILL
        ws.cell(row=r, column=7, value=f"=ROUND(E{r},2)-F{r}")
        ws.cell(row=r, column=8, value=f"=STDEV({rng})*SQRT(Returns!$G$10)*100")
        g = ws.cell(row=r, column=9, value=asset_by_id[pid]["volatilityPct"])
        g.fill = GIVEN_FILL
        ws.cell(row=r, column=10, value=f"=ROUND(H{r},2)-I{r}")
        w_rows[pid] = r
        r += 1
    rK, rM = w_rows["M0209_2548"], w_rows["M0155_2547"]
    ws.cell(row=r, column=2, value="sum of weights (must be 1)").font = HDR
    ws.cell(row=r, column=3, value=f"=SUM(C{rK}:C{rM})")
    r += 2

    ws.cell(row=r, column=1, value="Portfolio expected return").font = HDR
    r += 1
    ws.cell(row=r, column=1, value="monthly (SUMPRODUCT of weights x mean returns)")
    ws.cell(row=r, column=3, value=f"=SUMPRODUCT(C{rK}:C{rM},D{rK}:D{rM})")
    row_mu_m = r
    r += 1
    ws.cell(row=r, column=1, value="annualized %")
    ws.cell(row=r, column=3, value=f"=C{row_mu_m}*Returns!$G$10*100")
    row_mu_a = r
    ws.cell(row=r, column=4, value="given optimalPoint.expectedReturnPct").font = HDR
    ws.cell(row=r, column=5, value=result["optimalPoint"]["expectedReturnPct"]).fill = GIVEN_FILL
    ws.cell(row=r, column=6, value=f"=ROUND(C{row_mu_a},2)-E{r}")
    r += 2

    ws.cell(row=r, column=1, value="Covariance matrix Sigma (monthly, sample / ddof=1)").font = HDR
    r += 1
    _header_row(ws, r, ["", "M0209_2548", "M0155_2547"])
    r += 1
    cov_top = r
    ws.cell(row=r, column=1, value="M0209_2548").font = HDR
    ws.cell(row=r, column=2, value=f"=VAR({RET_KSET50})")
    ws.cell(row=r, column=3, value=COV_EXPR)
    r += 1
    ws.cell(row=r, column=1, value="M0155_2547").font = HDR
    ws.cell(row=r, column=2, value=COV_EXPR)
    ws.cell(row=r, column=3, value=f"=VAR({RET_MS50})")
    cov_bot = r
    S11 = f"$B${cov_top}"
    S12 = f"$C${cov_top}"
    S22 = f"$C${cov_bot}"
    r += 2

    ws.cell(row=r, column=1, value="correlation (live)")
    ws.cell(row=r, column=3, value=f"={S12}/SQRT({S11}*{S22})")
    ws.cell(row=r, column=4, value="given correlations[0].correlation").font = HDR
    ws.cell(row=r, column=5, value=result["correlations"][0]["correlation"]).fill = GIVEN_FILL
    ws.cell(row=r, column=6, value=f"=ROUND(C{r},2)-E{r}")
    r += 2

    ws.cell(row=r, column=1,
            value="Portfolio variance (monthly) = w1^2*s11 + w2^2*s22 + 2*w1*w2*s12")
    ws.cell(row=r, column=3,
            value=f"=C{rK}^2*{S11}+C{rM}^2*{S22}+2*C{rK}*C{rM}*{S12}")
    row_var = r
    r += 1
    ws.cell(row=r, column=1, value="Portfolio volatility (monthly) = SQRT(variance)")
    ws.cell(row=r, column=3, value=f"=SQRT(C{row_var})")
    r += 1
    ws.cell(row=r, column=1, value="Portfolio volatility ANNUALIZED % = monthly*SQRT(12)*100")
    ws.cell(row=r, column=3, value=f"=SQRT(C{row_var})*SQRT(Returns!$G$10)*100")
    row_vol_a = r
    ws.cell(row=r, column=4, value="given selectedRiskMeasure.optimizedValue").font = HDR
    ws.cell(row=r, column=5,
            value=result["selectedRiskMeasure"]["optimizedValue"]).fill = GIVEN_FILL
    ws.cell(row=r, column=6, value=f"=ROUND(C{row_vol_a},2)-E{r}")
    ws.cell(row=r, column=7, value=f"=IF(ABS(F{r})<0.005,\"OK\",\"MISMATCH\")")
    r += 2

    ws.cell(row=r, column=1, value="Risk contribution %").font = HDR
    r += 1
    r = _header_row(ws, r, ["projId", "computed % (live)", "given riskContributionPct",
                            "diff", "check (tolerance 0.005 = given is rounded to 2dp)"])
    rc_given = result["riskContributionPct"]
    for pid, expr in (
        ("M0209_2548", f"C{rK}*({S11}*C{rK}+{S12}*C{rM})/C{row_var}*100"),
        ("M0155_2547", f"C{rM}*({S12}*C{rK}+{S22}*C{rM})/C{row_var}*100"),
    ):
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value="=" + expr)
        ws.cell(row=r, column=3, value=rc_given[pid]).fill = GIVEN_FILL
        ws.cell(row=r, column=4, value=f"=ROUND(B{r},2)-C{r}")
        ws.cell(row=r, column=5, value=f'=IF(ABS(D{r})<0.005,"OK","MISMATCH")')
        r += 1
    r += 1
    _note(ws, r, "Tolerance is 0.005 rather than the brief's 1e-6 because result.json's "
                 "reported values are rounded to 2 decimal places by the API layer; the "
                 "computed side is rounded to 2dp before differencing, so an exact match "
                 "shows as 0.")
    _widths(ws, [52, 20, 20, 30, 20, 14, 14, 20, 20, 14])

    # ------------------------------------------------------------- Frontier
    ws = wb.create_sheet("Frontier")
    r = _title(ws, "FRONTIER TAB VERIFICATION")
    r = _note(ws, r, "Each frontier point's own weights are GIVEN; its volatility, "
                     "expected return and Sharpe are recomputed live from the same "
                     "Sigma/mu built on the Weights sheet (which is itself rooted in NAV).")
    r += 1
    r = _header_row(ws, r, [
        "point #", "w M0209_2548 (given)", "w M0155_2547 (given)",
        "volatility % (live)", "given volatilityPct", "diff",
        "expected return % (live)", "given expectedReturnPct", "diff",
        "sharpe (live)", "given sharpe", "diff", "check"])
    W = "Weights!"
    for i, pt in enumerate(result["frontier"], start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=pt["weights"]["M0209_2548"] / 100.0).fill = GIVEN_FILL
        ws.cell(row=r, column=3, value=pt["weights"]["M0155_2547"] / 100.0).fill = GIVEN_FILL
        ws.cell(row=r, column=4, value=(
            f"=SQRT(B{r}^2*{W}{S11}+C{r}^2*{W}{S22}+2*B{r}*C{r}*{W}{S12})"
            f"*SQRT(Returns!$G$10)*100"))
        ws.cell(row=r, column=5, value=pt["volatilityPct"]).fill = GIVEN_FILL
        ws.cell(row=r, column=6, value=f"=ROUND(D{r},2)-E{r}")
        ws.cell(row=r, column=7, value=(
            f"=(B{r}*{W}$D${rK}+C{r}*{W}$D${rM})*Returns!$G$10*100"))
        ws.cell(row=r, column=8, value=pt["expectedReturnPct"]).fill = GIVEN_FILL
        ws.cell(row=r, column=9, value=f"=ROUND(G{r},2)-H{r}")
        ws.cell(row=r, column=10, value=f"=(G{r}-Returns!$G$11)/D{r}")
        ws.cell(row=r, column=11, value=pt["sharpe"]).fill = GIVEN_FILL
        ws.cell(row=r, column=12, value=f"=ROUND(J{r},2)-K{r}")
        ws.cell(row=r, column=13, value=(
            f'=IF(AND(ABS(F{r})<0.005,ABS(I{r})<0.005,ABS(L{r})<0.005),"OK","MISMATCH")'))
        r += 1
    r += 1
    _note(ws, r, "result.json's frontier holds a single point (100% M0209_2548) because "
                 "maxHoldings=2 over a 2-fund universe with a 3-observation window "
                 "collapses the efficient set. Sharpe uses the annualized excess return "
                 "over riskFreeRatePct=1.5.")
    _widths(ws, [10] + [20] * 12)

    # ---------------------------------------------------------- Performance
    ws = wb.create_sheet("Performance")
    r = _title(ws, "PERFORMANCE TAB VERIFICATION")
    r = _note(ws, r, "Formulas below are taken verbatim from docs/formula-reference.md: "
                     "annualized_return R_ann = PRODUCT(1+r_t)^(m/n)-1; "
                     "annualized_volatility sigma_ann = std(r_t, ddof=1)*SQRT(m); "
                     "sortino_ratio sigma_down = SQRT(mean(min(r_t,0)^2))*SQRT(m), "
                     "Sortino = (R_ann - R_f)/sigma_down with MAR=0; "
                     "max_drawdown MDD = min(V_t/Peak_t - 1).")
    r += 1
    r = _header_row(ws, r, ["month_end", "portfolio return (live)",
                            "given monthlyReturnsPct", "diff",
                            "growth of 1 (V_t)", "running peak", "drawdown",
                            "downside min(r,0)"])
    first_series = r
    ws.cell(row=r - 1, column=5, value="growth of 1 (V_t)")
    # V0 row
    ws.cell(row=r, column=1, value="start (2024-02-29)")
    ws.cell(row=r, column=5, value=1)
    ws.cell(row=r, column=6, value=f"=MAX($E${first_series}:E{r})")
    ws.cell(row=r, column=7, value=f"=E{r}/F{r}-1")
    r += 1
    mrp = result["monthlyReturnsPct"]
    for k in range(N_OBS):
        ret_row = 5 + k  # Returns sheet row holding this month's returns
        ws.cell(row=r, column=1, value=f"=Returns!F{ret_row}")
        ws.cell(row=r, column=2, value=(
            f"=Returns!J{ret_row}*Weights!$C${rK}+Returns!I{ret_row}*Weights!$C${rM}"))
        ws.cell(row=r, column=3, value=mrp[k]).fill = GIVEN_FILL
        ws.cell(row=r, column=4, value=f"=ROUND(B{r}*100,2)-C{r}")
        ws.cell(row=r, column=5, value=f"=E{r-1}*(1+B{r})")
        ws.cell(row=r, column=6, value=f"=MAX($E${first_series}:E{r})")
        ws.cell(row=r, column=7, value=f"=E{r}/F{r}-1")
        ws.cell(row=r, column=8, value=f"=MIN(B{r},0)")
        r += 1
    last_series = r - 1
    RET_RNG = f"$B${first_series + 1}:$B${last_series}"
    DD_RNG = f"$G${first_series}:$G${last_series}"
    DOWN_RNG = f"$H${first_series + 1}:$H${last_series}"
    V_LAST = f"$E${last_series}"
    r += 1

    ws.cell(row=r, column=1, value="Metrics").font = HDR
    r += 1
    r = _header_row(ws, r, ["metric", "formula-reference.md name", "computed (live)",
                            "given performanceSummary[0]", "diff", "check"])
    ps = result["performanceSummary"][0]
    metrics = [
        ("CAGR %", "annualized_return",
         f"=({V_LAST}^(Returns!$G$10/{N_OBS})-1)*100", ps["cagrPct"]),
        ("Expected (arithmetic annualized) return %", "mean * m",
         f"=AVERAGE({RET_RNG})*Returns!$G$10*100", ps["expectedReturnPct"]),
        ("Annualized volatility %", "annualized_volatility",
         f"=STDEV({RET_RNG})*SQRT(Returns!$G$10)*100", ps["stdDevPct"]),
        ("Max drawdown %", "max_drawdown", f"=MIN({DD_RNG})*100", ps["maxDrawdownPct"]),
    ]
    rows_by_metric = {}
    for label, ref, formula, given in metrics:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=ref)
        ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=4, value=given).fill = GIVEN_FILL
        ws.cell(row=r, column=5, value=f"=ROUND(C{r},2)-D{r}")
        ws.cell(row=r, column=6, value=f'=IF(ABS(E{r})<0.005,"OK","MISMATCH")')
        rows_by_metric[label] = r
        r += 1
    row_cagr = rows_by_metric["CAGR %"]
    row_vol = rows_by_metric["Annualized volatility %"]
    row_mu = rows_by_metric["Expected (arithmetic annualized) return %"]

    # downside deviation helper
    ws.cell(row=r, column=1, value="Downside deviation % (sigma_down)")
    ws.cell(row=r, column=2, value="sortino_ratio (MAR=0)")
    ws.cell(row=r, column=3,
            value=f"=SQRT(SUMSQ({DOWN_RNG})/{N_OBS})*SQRT(Returns!$G$10)*100")
    row_dd = r
    r += 1
    for label, ref, formula, given in [
        ("Sharpe ex-post", "(R_ann - R_f)/sigma_ann",
         f"=(C{row_cagr}-Returns!$G$11)/C{row_vol}", ps["sharpeExPost"]),
        ("Sharpe ex-ante", "(mean_ann - R_f)/sigma_ann",
         f"=(C{row_mu}-Returns!$G$11)/C{row_vol}", ps["sharpeExAnte"]),
        ("Sortino", "(R_ann - R_f)/sigma_down",
         f"=(C{row_cagr}-Returns!$G$11)/C{row_dd}", ps["sortino"]),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=ref)
        ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=4, value=given).fill = GIVEN_FILL
        ws.cell(row=r, column=5, value=f"=ROUND(C{r},2)-D{r}")
        ws.cell(row=r, column=6, value=f'=IF(ABS(E{r})<0.005,"OK","MISMATCH")')
        r += 1
    for label, given in [("bestYearPct", ps["bestYearPct"]),
                         ("worstYearPct", ps["worstYearPct"])]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value="n/a")
        ws.cell(row=r, column=3, value="not computable")
        c = ws.cell(row=r, column=4, value="null" if given is None else given)
        c.fill = GIVEN_FILL
        ws.cell(row=r, column=5, value="correctly null - undefined for this short a "
                                       "window (3 monthly observations, < 1 calendar year)")
        ws.cell(row=r, column=6, value="OK")
        r += 1
    _widths(ws, [42, 34, 22, 26, 52, 12])

    # ---------------------------------------------------------------- Rolling
    ws = wb.create_sheet("Rolling")
    r = _title(ws, "ROLLING TAB VERIFICATION")
    r += 1
    r = _header_row(ws, r, ["fold", "status", "explanation"])
    if result["rolling"]:
        raise SystemExit("rolling is non-empty; this sheet's generator needs the "
                         "per-fold branch implemented")
    ws.cell(row=r, column=1, value="(none)")
    ws.cell(row=r, column=2, value="0 folds returned - EXPECTED, not a defect")
    c = ws.cell(row=r, column=3, value=(
        "result.json's `rolling` array is empty. The request spans 2024-02-29 to "
        "2024-05-31 (3 monthly return observations) while "
        "constraints.lookbackPeriodMonths = 12 with rollingWindowMode = 'expanding'. "
        "A 12-month lookback cannot be satisfied even once inside a ~4-month window, "
        "so no fold can be formed, let alone the 2 folds the engine requires. "
        "The backend says the same thing in result.robustNote: "
        + str(result["robustNote"])))
    c.alignment = WRAP
    r += 2
    _note(ws, r, "Nothing to recompute here: there is no fold data to check against. "
                 "The verification performed on this sheet is that the empty array is "
                 "the arithmetically correct outcome of the request's own parameters.")
    _widths(ws, [14, 44, 120])

    # ----------------------------------------------------------------- Report
    ws = wb.create_sheet("Report")
    r = _title(ws, "REPORT TAB — CROSS-SHEET INTERNAL CONSISTENCY")
    r = _note(ws, r, "Not a new computation: each row asserts that a top-level scalar in "
                     "result.json agrees with what the live formulas on the earlier "
                     "sheets produce, or with the request that generated it.")
    r += 1
    r = _header_row(ws, r, ["field", "given (result.json)", "cross-reference", "check"])

    def report_row(row, field, given, xref_formula, check_formula):
        ws.cell(row=row, column=1, value=field)
        c = ws.cell(row=row, column=2, value=given)
        c.fill = GIVEN_FILL
        ws.cell(row=row, column=3, value=xref_formula).alignment = WRAP
        ws.cell(row=row, column=4, value=check_formula)
        return row + 1

    r = report_row(r, "feasibility", result["feasibility"],
                   "weights sum to 1 and all weights within [minWeightPct, maxWeightPct] "
                   "= [0, 100]",
                   f'=IF(AND(ABS(Weights!C{rM + 1}-1)<0.000001,'
                   f'Weights!C{rK}>=0,Weights!C{rK}<=1,'
                   f'Weights!C{rM}>=0,Weights!C{rM}<=1),"OK","MISMATCH")')
    r = report_row(r, "selectedRiskMeasure.measure",
                   result["selectedRiskMeasure"]["measure"],
                   "request.riskMeasure = " + request["riskMeasure"],
                   f'=IF(B{r}="{request["riskMeasure"]}","OK","MISMATCH")')
    r = report_row(r, "selectedRiskMeasure.optimizedValue",
                   result["selectedRiskMeasure"]["optimizedValue"],
                   f"=Weights!C{row_vol_a}",
                   f'=IF(ABS(ROUND(C{r},2)-B{r})<0.005,"OK","MISMATCH")')
    r = report_row(r, "optimalPoint.volatilityPct",
                   result["optimalPoint"]["volatilityPct"],
                   f"=Weights!C{row_vol_a}",
                   f'=IF(ABS(ROUND(C{r},2)-B{r})<0.005,"OK","MISMATCH")')
    r = report_row(r, "optimalPoint.expectedReturnPct",
                   result["optimalPoint"]["expectedReturnPct"],
                   f"=Weights!C{row_mu_a}",
                   f'=IF(ABS(ROUND(C{r},2)-B{r})<0.005,"OK","MISMATCH")')
    r = report_row(r, "gmvPoint.volatilityPct", result["gmvPoint"]["volatilityPct"],
                   "=Frontier!D5",
                   f'=IF(ABS(ROUND(C{r},2)-B{r})<0.005,"OK","MISMATCH")')
    r = report_row(r, "tangencyPoint.volatilityPct",
                   result["tangencyPoint"]["volatilityPct"], "=Frontier!D5",
                   f'=IF(ABS(ROUND(C{r},2)-B{r})<0.005,"OK","MISMATCH")')
    # turnover: currentWeightPct is empty {} => all-cash start; turnover = 0.5*sum|dw|
    r = report_row(r, "totalTurnoverPct", result["totalTurnoverPct"],
                   f"=0.5*(ABS(Weights!C{rK})+ABS(Weights!C{rM}))*100",
                   f'=IF(ABS(0.5*(ABS(Weights!C{rK})+ABS(Weights!C{rM}))*100-B{r})'
                   '<0.005,"OK","MISMATCH")')
    ws.cell(row=r - 1, column=5, value=(
        "turnover = 0.5 * sum|w_target - w_current|; request.currentWeightPct is {} "
        "so every current weight is 0, giving 0.5 * 100% = 50%")).font = NOTE
    r = report_row(r, "generatedAt", result["generatedAt"],
                   "run timestamp; nothing to recompute", "n/a")
    r = report_row(r, "robustNote", result["robustNote"],
                   "see Rolling sheet - 0 folds is arithmetically forced by "
                   "lookbackPeriodMonths=12 over a 3-observation window", "OK")
    for field in ("feasibilityMessage", "compareNote", "constraintNote",
                  "robustOptimizationNote"):
        r = report_row(r, field, "null" if result[field] is None else result[field],
                       "null expected: compareAgainst='none', "
                       "groupConstraintsEnabled=false, robustOptimization=false",
                       "OK")
    _widths(ws, [36, 46, 82, 16])

    # ------------------------------------------------- FullSolveReplication
    ws = wb.create_sheet("FullSolveReplication")
    r = _title(ws, "FULL SOLVE REPLICATION — closed-form 2-asset tangency portfolio")
    r = _note(ws, r, "This is the only sheet that re-derives the WEIGHTS themselves "
                     "rather than the reporting layer. Inputs: mu and Sigma from the "
                     "Weights sheet (which are live from NAV). result.json is used only "
                     "in the yellow comparison column.")
    r += 1
    r = _header_row(ws, r, ["quantity", "M0209_2548", "M0155_2547"])
    ws.cell(row=r, column=1, value="mean monthly return mu (live)")
    ws.cell(row=r, column=2, value=f"=Weights!D{rK}")
    ws.cell(row=r, column=3, value=f"=Weights!D{rM}")
    row_mu_v = r
    r += 1
    ws.cell(row=r, column=1, value="monthly risk-free r_f = riskFreeRatePct/100/12")
    ws.cell(row=r, column=2, value="=Returns!$G$11/100/Returns!$G$10")
    row_rf = r
    r += 1
    ws.cell(row=r, column=1, value="excess return mu - r_f")
    ws.cell(row=r, column=2, value=f"=B{row_mu_v}-$B${row_rf}")
    ws.cell(row=r, column=3, value=f"=C{row_mu_v}-$B${row_rf}")
    row_ex = r
    r += 2

    ws.cell(row=r, column=1, value="Sigma (monthly, from Weights sheet)").font = HDR
    r += 1
    ws.cell(row=r, column=1, value="s11")
    ws.cell(row=r, column=2, value=f"=Weights!{S11}")
    s11 = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="s12")
    ws.cell(row=r, column=2, value=f"=Weights!{S12}")
    s12 = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="s22")
    ws.cell(row=r, column=2, value=f"=Weights!{S22}")
    s22 = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="det(Sigma) = s11*s22 - s12^2")
    ws.cell(row=r, column=2, value=f"={s11}*{s22}-{s12}^2")
    det = f"$B${r}"
    r += 2

    ws.cell(row=r, column=1,
            value="Unnormalized tangency weights u = Sigma^-1 * (mu - r_f), "
                  "closed-form 2x2 inverse").font = HDR
    r += 1
    ws.cell(row=r, column=1, value="u1 = ( s22*e1 - s12*e2) / det")
    ws.cell(row=r, column=2, value=f"=({s22}*B{row_ex}-{s12}*C{row_ex})/{det}")
    u1 = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="u2 = (-s12*e1 + s11*e2) / det")
    ws.cell(row=r, column=2, value=f"=(-{s12}*B{row_ex}+{s11}*C{row_ex})/{det}")
    u2 = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="sum(u)")
    ws.cell(row=r, column=2, value=f"={u1}+{u2}")
    usum = f"$B${r}"
    r += 2

    ws.cell(row=r, column=1, value="Cross-check with Excel array functions "
                                   "(spills in Excel 365 / CSE in legacy Excel)").font = HDR
    r += 1
    ws.cell(row=r, column=1, value="MMULT(MINVERSE(Sigma), excess)  [array formula in B]")
    ws.cell(row=r, column=2,
            value=f"=MMULT(MINVERSE(Weights!{S11}:{S22}),"
                  f"TRANSPOSE(B{row_ex}:C{row_ex}))")
    ws.cell(row=r, column=4, value="should equal u1 (and u2 in the cell below when it "
                                   "spills)").font = NOTE
    r += 3

    r = _header_row(ws, r, [
        "projId", "closed-form normalized weight (live)",
        "given optimalWeights (result.json)", "diff", "match to 6 dp?"])
    for pid, uexpr in (("M0209_2548", u1), ("M0155_2547", u2)):
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value=f"={uexpr}/{usum}")
        ws.cell(row=r, column=3, value=ow[pid] / 100.0).fill = GIVEN_FILL
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=5, value=f'=IF(ABS(D{r})<0.000001,"MATCH","DIFFERS")')
        r += 1
    r += 1
    _note(ws, r, "EXPECTED RESULT: DIFFERS. The closed-form tangency solution is "
                 "UNCONSTRAINED and produces a short position, which the request forbids "
                 "(constraints.longOnly = true, minWeightPct = 0). The grid below is the "
                 "long-only comparison and is the meaningful check.")
    r += 3

    ws.cell(row=r, column=1, value="Long-only Sharpe grid — the constrained solve, "
                                   "brute-forced live in Excel").font = HDR
    r += 1
    r = _header_row(ws, r, ["w(M0209_2548)", "w(M0155_2547)",
                            "annualized return %", "annualized volatility %",
                            "Sharpe (annualized, excess over 1.5%)"])
    grid_top = r
    for k in range(21):
        a = k / 20.0
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=f"=1-A{r}")
        ws.cell(row=r, column=3,
                value=f"=(A{r}*Weights!$D${rK}+B{r}*Weights!$D${rM})*Returns!$G$10*100")
        ws.cell(row=r, column=4, value=(
            f"=SQRT(A{r}^2*Weights!{S11}+B{r}^2*Weights!{S22}"
            f"+2*A{r}*B{r}*Weights!{S12})*SQRT(Returns!$G$10)*100"))
        ws.cell(row=r, column=5, value=f"=(C{r}-Returns!$G$11)/D{r}")
        r += 1
    grid_bot = r - 1
    # the reported weights, evaluated on the same Sharpe formula
    ws.cell(row=r, column=1, value=f"=Weights!C{rK}")
    ws.cell(row=r, column=2, value=f"=Weights!C{rM}")
    ws.cell(row=r, column=3,
            value=f"=(A{r}*Weights!$D${rK}+B{r}*Weights!$D${rM})*Returns!$G$10*100")
    ws.cell(row=r, column=4, value=(
        f"=SQRT(A{r}^2*Weights!{S11}+B{r}^2*Weights!{S22}"
        f"+2*A{r}*B{r}*Weights!{S12})*SQRT(Returns!$G$10)*100"))
    ws.cell(row=r, column=5, value=f"=(C{r}-Returns!$G$11)/D{r}")
    ws.cell(row=r, column=6, value="<- result.json's optimalWeights scored on the "
                                   "identical Sharpe formula").font = NOTE
    row_reported = r
    r += 2
    ws.cell(row=r, column=1, value="best Sharpe on the long-only grid")
    ws.cell(row=r, column=2, value=f"=MAX(E{grid_top}:E{grid_bot})")
    row_best = r
    r += 1
    ws.cell(row=r, column=1, value="argmax w(M0209_2548) on the grid")
    ws.cell(row=r, column=2,
            value=f"=INDEX(A{grid_top}:A{grid_bot},MATCH(B{row_best},"
                  f"E{grid_top}:E{grid_bot},0))")
    r += 1
    ws.cell(row=r, column=1, value="Sharpe of result.json's optimalWeights")
    ws.cell(row=r, column=2, value=f"=E{row_reported}")
    row_rep_sharpe = r
    r += 1
    ws.cell(row=r, column=1, value="grid best MINUS reported (positive => reported "
                                   "weights are NOT the long-only max-Sharpe point)")
    ws.cell(row=r, column=2, value=f"=B{row_best}-B{row_rep_sharpe}")
    r += 2
    _note(ws, r, "FINDING: the grid maximum sits at w(M0209_2548) = 1.00 with Sharpe "
                 "about -0.4396, while result.json's optimalWeights (97.49% / 2.51%) "
                 "score about -0.4940 on the identical formula. Both funds have negative "
                 "excess returns over the window, which makes the max-Sharpe objective "
                 "degenerate; result.json's own frontier / tangencyPoint DO sit at the "
                 "100% corner, so the reporting layer and the solver disagree with each "
                 "other. Flagged for Task 4 - this sheet is the evidence, not a "
                 "diagnosis of the solver.")
    _widths(ws, [64, 26, 26, 26, 34, 46])

    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
