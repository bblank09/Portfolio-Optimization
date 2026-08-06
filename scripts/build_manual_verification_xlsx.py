"""Builds the manual verification workbook for Phase 4.4.

Every calculation cell is a real Excel formula, independently expressed in
Excel's own function language -- not a copy of the Python engine's output.
Blue = hardcoded input, black = formula, green = cross-sheet reference.
Source data: docs/verification/{4.3-raw-nav-inputs.csv, 4.3b-result.json}.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=14)
PCT6 = "0.000000%"
NUM6 = "0.000000"
MONEY = "#,##0.0000"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROOT = Path(__file__).resolve().parents[1]
RESULT_PATH = ROOT / "docs" / "verification" / "4.3b-result.json"
OUT_PATH = ROOT / "out" / "formula-manual-verification-2026-08-02.xlsx"

result = json.loads(RESULT_PATH.read_text(encoding="utf-8"))

wb = Workbook()

# ---------------------------------------------------------------- Inputs --
inp = wb.active
inp.title = "Inputs"
inp["B2"] = "Run 2 — cashflow + rebalancing manual verification"
inp["B2"].font = TITLE_FONT
inp["B3"] = f"run_id: {result['run_id']}  |  generated: {result['created_at']}"

inp["B5"] = "Portfolio parameters"
inp["B5"].font = BOLD
params = [
    ("Initial capital", 100000, None),
    ("Target weight K-SET50", 0.60, "0%"),
    ("Target weight M-S50", 0.40, "0%"),
    ("Risk-free rate (annual)", 0.02, "0%"),
    ("Periods per year (m)", 12, None),
    ("Monthly contribution", 2000, None),
    ("Transaction cost (bps)", 5, None),
    ("Slippage (bps)", 5, None),
]
row = 6
param_cells = {}
for label, value, fmt in params:
    inp[f"B{row}"] = label
    inp[f"C{row}"] = value
    inp[f"C{row}"].font = BLUE
    if fmt:
        inp[f"C{row}"].number_format = fmt
    param_cells[label] = f"Inputs!$C${row}"
    row += 1
inp[f"B{row}"] = "Cost rate (txn + slippage, decimal)"
inp[f"C{row}"] = f"=({param_cells['Transaction cost (bps)']}+{param_cells['Slippage (bps)']})/10000"
inp[f"C{row}"].font = BLACK
inp[f"C{row}"].number_format = "0.000000"
param_cells["Cost rate"] = f"Inputs!$C${row}"
row += 2

inp[f"B{row}"] = "Raw month-end NAV per unit (source: docs/verification/4.3-raw-nav-inputs.csv)"
inp[f"B{row}"].font = BOLD
row += 1
nav_header_row = row
dates = ["2023-01-31", "2023-02-28", "2023-03-31", "2023-04-30", "2023-05-31"]
inp.cell(row=row, column=2, value="Date")
for i, d in enumerate(dates):
    inp.cell(row=row, column=3 + i, value=d).font = BLUE
row += 1
nav_kset_row = row
nav_values = {
    "K-SET50": [34.6300, 33.8033, 34.2342, 32.6194, 32.7233],
    "M-S50": [23.2908, 22.7281, 23.0008, 21.9503, 21.4319],
}
inp.cell(row=row, column=2, value="NAV K-SET50")
for i, v in enumerate(nav_values["K-SET50"]):
    c = inp.cell(row=row, column=3 + i, value=v)
    c.font = BLUE
    c.number_format = MONEY
row += 1
nav_ms50_row = row
inp.cell(row=row, column=2, value="NAV M-S50")
for i, v in enumerate(nav_values["M-S50"]):
    c = inp.cell(row=row, column=3 + i, value=v)
    c.font = BLUE
    c.number_format = MONEY
row += 1

nav_kset_ref = lambda col: f"Inputs!{get_column_letter(col)}${nav_kset_row}"
nav_ms50_ref = lambda col: f"Inputs!{get_column_letter(col)}${nav_ms50_row}"
date_ref = lambda col: f"Inputs!{get_column_letter(col)}${nav_header_row}"

inp.column_dimensions["B"].width = 34
for col in range(3, 8):
    inp.column_dimensions[get_column_letter(col)].width = 14

# ------------------------------------------------------------ Simulation --
sim = wb.create_sheet("Simulation")
sim["B2"] = "Monthly portfolio simulation ledger (independent Excel formulas)"
sim["B2"].font = TITLE_FONT
sim["B3"] = "Replicates backend/app/engine/backtest.py's per-period loop exactly, one column per period."

col0 = 3  # column C = t0 (2023-01-31)
periods = list(range(5))  # t0..t4

def col_for(t):
    return col0 + t

rows = {}
r = 5
def add_row(key, label):
    global r
    sim.cell(row=r, column=2, value=label).font = BOLD if key in ("EndingValue", "PeriodPerformance") else Font()
    rows[key] = r
    r += 1

add_row("Date", "Date")
add_row("NAV_KSET", "NAV K-SET50")
add_row("NAV_MS50", "NAV M-S50")
add_row("Ret_KSET", "Return K-SET50 (r_t)")
add_row("Ret_MS50", "Return M-S50 (r_t)")
add_row("StartValue", "Starting portfolio value")
add_row("Val_KSET_ret", "K-SET50 value after return")
add_row("Val_MS50_ret", "M-S50 value after return")
add_row("Cashflow", "Cashflow applied (end of period)")
add_row("Val_KSET_cf", "K-SET50 value after cashflow")
add_row("Val_MS50_cf", "M-S50 value after cashflow")
add_row("TotalPreRebal", "Total value before rebalance")
add_row("Target_KSET", "Target K-SET50 value (60%)")
add_row("Target_MS50", "Target M-S50 value (40%)")
add_row("MoneyTurnover", "Money turnover = sum|target-current|/2")
add_row("TurnoverRatio", "Turnover ratio = turnover/total")
add_row("Cost", "Rebalance cost = turnover * cost_rate")
add_row("Val_KSET_final", "K-SET50 value after cost")
add_row("Val_MS50_final", "M-S50 value after cost")
add_row("EndingValue", "Ending portfolio value")
add_row("PeriodPerformance", "Period performance (cashflow-neutral)")

for t in periods:
    c = col_for(t)
    L = get_column_letter(c)
    sim.cell(row=rows["Date"], column=c, value=f"={date_ref(3 + t)}")
    sim.cell(row=rows["NAV_KSET"], column=c, value=f"={nav_kset_ref(3 + t)}")
    sim.cell(row=rows["NAV_MS50"], column=c, value=f"={nav_ms50_ref(3 + t)}")

    if t == 0:
        sim.cell(row=rows["Ret_KSET"], column=c, value=None)
        sim.cell(row=rows["Ret_MS50"], column=c, value=None)
        sim.cell(row=rows["StartValue"], column=c, value=None)
        sim.cell(row=rows["Val_KSET_ret"], column=c, value=f"={param_cells['Initial capital']}*{param_cells['Target weight K-SET50']}")
        sim.cell(row=rows["Val_MS50_ret"], column=c, value=f"={param_cells['Initial capital']}*{param_cells['Target weight M-S50']}")
        sim.cell(row=rows["Cashflow"], column=c, value=0)
        sim.cell(row=rows["Cashflow"], column=c).font = BLUE
    else:
        prev = get_column_letter(c - 1)
        sim.cell(row=rows["Ret_KSET"], column=c, value=f"={L}{rows['NAV_KSET']}/{prev}{rows['NAV_KSET']}-1")
        sim.cell(row=rows["Ret_MS50"], column=c, value=f"={L}{rows['NAV_MS50']}/{prev}{rows['NAV_MS50']}-1")
        sim.cell(row=rows["StartValue"], column=c, value=f"={prev}{rows['EndingValue']}")
        sim.cell(row=rows["Val_KSET_ret"], column=c, value=f"={prev}{rows['Val_KSET_final']}*(1+{L}{rows['Ret_KSET']})")
        sim.cell(row=rows["Val_MS50_ret"], column=c, value=f"={prev}{rows['Val_MS50_final']}*(1+{L}{rows['Ret_MS50']})")
        sim.cell(row=rows["Cashflow"], column=c, value=f"={param_cells['Monthly contribution']}")

    sim.cell(row=rows["Val_KSET_cf"], column=c, value=f"={L}{rows['Val_KSET_ret']}+{param_cells['Target weight K-SET50']}*{L}{rows['Cashflow']}")
    sim.cell(row=rows["Val_MS50_cf"], column=c, value=f"={L}{rows['Val_MS50_ret']}+{param_cells['Target weight M-S50']}*{L}{rows['Cashflow']}")
    sim.cell(row=rows["TotalPreRebal"], column=c, value=f"={L}{rows['Val_KSET_cf']}+{L}{rows['Val_MS50_cf']}")
    sim.cell(row=rows["Target_KSET"], column=c, value=f"={L}{rows['TotalPreRebal']}*{param_cells['Target weight K-SET50']}")
    sim.cell(row=rows["Target_MS50"], column=c, value=f"={L}{rows['TotalPreRebal']}*{param_cells['Target weight M-S50']}")
    sim.cell(row=rows["MoneyTurnover"], column=c, value=f"=(ABS({L}{rows['Target_KSET']}-{L}{rows['Val_KSET_cf']})+ABS({L}{rows['Target_MS50']}-{L}{rows['Val_MS50_cf']}))/2")
    sim.cell(row=rows["TurnoverRatio"], column=c, value=f"={L}{rows['MoneyTurnover']}/{L}{rows['TotalPreRebal']}")
    sim.cell(row=rows["Cost"], column=c, value=f"={L}{rows['MoneyTurnover']}*{param_cells['Cost rate']}")
    sim.cell(row=rows["Val_KSET_final"], column=c, value=f"={L}{rows['Target_KSET']}-({L}{rows['Target_KSET']}/{L}{rows['TotalPreRebal']})*{L}{rows['Cost']}")
    sim.cell(row=rows["Val_MS50_final"], column=c, value=f"={L}{rows['Target_MS50']}-({L}{rows['Target_MS50']}/{L}{rows['TotalPreRebal']})*{L}{rows['Cost']}")
    sim.cell(row=rows["EndingValue"], column=c, value=f"={L}{rows['Val_KSET_final']}+{L}{rows['Val_MS50_final']}")

    if t == 0:
        sim.cell(row=rows["PeriodPerformance"], column=c, value=None)
    else:
        sim.cell(row=rows["PeriodPerformance"], column=c, value=f"=({L}{rows['EndingValue']}-{L}{rows['Cashflow']})/{L}{rows['StartValue']}-1")

    for key in rows:
        cell = sim.cell(row=rows[key], column=c)
        if cell.value is None:
            continue
        if key == "Cashflow" and t == 0:
            cell.font = BLUE
        elif key == "Date":
            cell.font = GREEN
        else:
            cell.font = BLACK
        if key in ("Ret_KSET", "Ret_MS50", "TurnoverRatio", "PeriodPerformance"):
            cell.number_format = PCT6
        elif key != "Date":
            cell.number_format = MONEY

sim.column_dimensions["B"].width = 34
for t in periods:
    sim.column_dimensions[get_column_letter(col_for(t))].width = 15

PP = f"Simulation!{get_column_letter(col_for(1))}{rows['PeriodPerformance']}:{get_column_letter(col_for(4))}{rows['PeriodPerformance']}"
RET_KSET = f"Simulation!{get_column_letter(col_for(1))}{rows['Ret_KSET']}:{get_column_letter(col_for(4))}{rows['Ret_KSET']}"
RET_MS50 = f"Simulation!{get_column_letter(col_for(1))}{rows['Ret_MS50']}:{get_column_letter(col_for(4))}{rows['Ret_MS50']}"
EQUITY = f"Simulation!{get_column_letter(col_for(0))}{rows['EndingValue']}:{get_column_letter(col_for(4))}{rows['EndingValue']}"
CASHFLOW_1_4 = f"Simulation!{get_column_letter(col_for(1))}{rows['Cashflow']}:{get_column_letter(col_for(4))}{rows['Cashflow']}"
TURNOVER_1_4 = f"Simulation!{get_column_letter(col_for(1))}{rows['MoneyTurnover']}:{get_column_letter(col_for(4))}{rows['MoneyTurnover']}"
COST_1_4 = f"Simulation!{get_column_letter(col_for(1))}{rows['Cost']}:{get_column_letter(col_for(4))}{rows['Cost']}"

# --------------------------------------------------------------- Metrics --
met = wb.create_sheet("Metrics")
met["B2"] = "Metrics — independent Excel formulas (see docs/formula-reference.md for the cited definitions)"
met["B2"].font = TITLE_FONT

def prod1p(rng):
    # PRODUCT(1+range) needs array-context (CSE) that isn't guaranteed
    # outside Excel 365's implicit spilling; EXP(SUMPRODUCT(LN(...))) is a
    # classic-Excel-compatible way to compute the same geometric product
    # without requiring an array-entered formula.
    return f"EXP(SUMPRODUCT(LN(1+{rng})))"

def stdev_s(rng):
    # STDEV.S is a post-2007 function name; openpyxl writes formulas as
    # plain strings, so without an _xlfn. prefix Excel/LibreOffice treat it
    # as an unrecognized name (#NAME?). STDEV (pre-2010) is the identical
    # sample (ddof=1) standard deviation under the old name.
    return f"STDEV({rng})"

def var_s(rng):
    return f"VAR({rng})"

def percentile_inc(rng, p):
    # PERCENTILE (pre-2010) is exactly PERCENTILE.INC under the old name.
    return f"PERCENTILE({rng},{p})"

def covariance_s(rng_x, rng_y):
    # No pre-2010 sample-covariance function exists (old COVAR() is
    # population covariance, dividing by n not n-1), so compute it directly:
    # cov_sample(X,Y) = sum((X-mean(X))*(Y-mean(Y))) / (n-1).
    return f"SUMPRODUCT(({rng_x}-AVERAGE({rng_x}))*({rng_y}-AVERAGE({rng_y})))/(COUNT({rng_x})-1)"

def stdev_diff(rng_x, rng_y):
    # STDEV.S(x-y) needs elementwise (x_i - y_i) inside a plain STDEV() call,
    # which -- like PRODUCT(1+range) -- needs CSE array-entry that a plain
    # formula string doesn't get. SUMPRODUCT is inherently array-aware, so
    # express the same sample-stdev-of-differences directly through it:
    # sqrt( sum((d_i - mean(d))^2) / (n-1) ), d_i = x_i - y_i.
    mean_diff = f"(AVERAGE({rng_x})-AVERAGE({rng_y}))"
    return f"SQRT(SUMPRODUCT(({rng_x}-{rng_y}-{mean_diff})^2)/(COUNT({rng_x})-1))"

mrow = 4
def m(label, formula, fmt=NUM6):
    global mrow
    met.cell(row=mrow, column=2, value=label)
    c = met.cell(row=mrow, column=3, value=formula)
    c.font = BLACK
    c.number_format = fmt
    ref = f"Metrics!$C${mrow}"
    mrow += 1
    return ref

met.cell(row=mrow, column=2, value="Return & risk").font = BOLD
mrow += 1
n_ref = m("n (period count)", f"=COUNT({PP})", "0")
twrr_ref = m("TWRR = Π(1+r_t) − 1", f"={prod1p(PP)}-1", PCT6)
cagr_ref = m("TWRR CAGR = (1+TWRR)^(m/n) − 1", f"=(1+{twrr_ref})^({param_cells['Periods per year (m)']}/{n_ref})-1", PCT6)
vol_ref = m("Volatility = STDEV.S(r_t)·√m", f"={stdev_s(PP)}*SQRT({param_cells['Periods per year (m)']})", PCT6)
sharpe_ref = m("Sharpe = (CAGR−Rf)/Vol", f"=({cagr_ref}-{param_cells['Risk-free rate (annual)']})/{vol_ref}", NUM6)

# Downside deviation helper row (array-safe: one cell per period via SUMPRODUCT)
downside_ref = m("Downside deviation = √(mean(min(r_t,0)²))·√m", f"=SQRT(SUMPRODUCT((({PP}<0)*{PP})^2)/{n_ref})*SQRT({param_cells['Periods per year (m)']})", PCT6)
sortino_ref = m("Sortino = (CAGR−Rf)/DownsideDev", f"=({cagr_ref}-{param_cells['Risk-free rate (annual)']})/{downside_ref}", NUM6)

var95_ref = m("VaR 95% = max(0, −PERCENTILE.INC(r_t,0.05))", f"=MAX(0,-{percentile_inc(PP, 0.05)})", PCT6)
var99_ref = m("VaR 99% = max(0, −PERCENTILE.INC(r_t,0.01))", f"=MAX(0,-{percentile_inc(PP, 0.01)})", PCT6)

# Max drawdown: for each of the 5 equity points, runningPeak_t = MAX(V_0..V_t);
# drawdown_t = V_t/runningPeak_t - 1; MDD = MIN(drawdown_t).
eq_cols = [get_column_letter(col_for(t)) for t in periods]
eq_row = rows["EndingValue"]
dd_terms = []
for L in eq_cols:
    peak_range = f"Simulation!${eq_cols[0]}${eq_row}:${L}${eq_row}"
    dd_terms.append(f"(Simulation!${L}${eq_row}/MAX({peak_range})-1)")
mdd_ref = m("Max drawdown = min(V_t/runningPeak_t − 1)", "=MIN(" + ",".join(dd_terms) + ")", PCT6)

calmar_ref = m("Calmar = CAGR / |MDD|", f"=IFERROR({cagr_ref}/ABS({mdd_ref}),\"n/a\")", NUM6)

met.cell(row=mrow, column=2, value="Money-weighted return (IRR)").font = BOLD
mrow += 1
irr_cf_row = mrow
met.cell(row=mrow, column=2, value="Investor cashflow (monthly-spaced, for IRR())")
irr_cols = []
for t in periods:
    L = get_column_letter(3 + t)
    cell = met.cell(row=mrow, column=3 + t)
    if t == 0:
        cell.value = f"=-{param_cells['Initial capital']}"
    elif t < 4:
        cell.value = f"=-Simulation!{get_column_letter(col_for(t))}{rows['Cashflow']}"
    else:
        cell.value = f"=-Simulation!{get_column_letter(col_for(t))}{rows['Cashflow']}+Simulation!{get_column_letter(col_for(t))}{rows['EndingValue']}"
    cell.font = BLACK
    cell.number_format = MONEY
    irr_cols.append(f"{get_column_letter(3+t)}{mrow}")
mrow += 1
irr_range = f"Metrics!$C${irr_cf_row}:$G${irr_cf_row}"
monthly_irr_ref = m("Monthly IRR = IRR(cashflow array)", f"=IRR({irr_range})", "0.000000%")
irr_ref = m("Annual IRR = (1+monthlyIRR)^m − 1", f"=(1+{monthly_irr_ref})^{param_cells['Periods per year (m)']}-1", PCT6)

met.cell(row=mrow, column=2, value="Benchmark-relative (benchmark = K-SET50)").font = BOLD
mrow += 1
bench_cagr_ref = m("Benchmark CAGR = Π(1+R_b,t)^(m/n) − 1", f"=({prod1p(RET_KSET)})^({param_cells['Periods per year (m)']}/{n_ref})-1", PCT6)
beta_ref = m("Beta = COVARIANCE.S(port,bench)/VAR.S(bench)", f"=({covariance_s(PP, RET_KSET)})/{var_s(RET_KSET)}", NUM6)
alpha_ref = m("Alpha = CAGR − [Rf + Beta·(BenchCAGR−Rf)]", f"={cagr_ref}-({param_cells['Risk-free rate (annual)']}+{beta_ref}*({bench_cagr_ref}-{param_cells['Risk-free rate (annual)']}))", PCT6)
te_ref = m("Tracking error = STDEV.S(port−bench)·√m", f"={stdev_diff(PP, RET_KSET)}*SQRT({param_cells['Periods per year (m)']})", PCT6)
ir_ref = m("Information ratio = (CAGR−BenchCAGR)/TE", f"=({cagr_ref}-{bench_cagr_ref})/{te_ref}", NUM6)
corr_bench_ref = m("Correlation (portfolio, benchmark)", f"=CORREL({PP},{RET_KSET})", NUM6)
excess_ref = m("Benchmark excess return = TWRR_port − TWRR_bench", f"={twrr_ref}-({prod1p(RET_KSET)}-1)", PCT6)

met.cell(row=mrow, column=2, value="Asset-level and diversification").font = BOLD
mrow += 1
kset_cagr_ref = m("K-SET50 CAGR", f"=({prod1p(RET_KSET)})^({param_cells['Periods per year (m)']}/{n_ref})-1", PCT6)
kset_vol_ref = m("K-SET50 volatility", f"={stdev_s(RET_KSET)}*SQRT({param_cells['Periods per year (m)']})", PCT6)
ms50_cagr_ref = m("M-S50 CAGR", f"=({prod1p(RET_MS50)})^({param_cells['Periods per year (m)']}/{n_ref})-1", PCT6)
ms50_vol_ref = m("M-S50 volatility", f"={stdev_s(RET_MS50)}*SQRT({param_cells['Periods per year (m)']})", PCT6)
asset_corr_ref = m("Correlation (K-SET50, M-S50) — Pearson, pairwise", f"=CORREL({RET_KSET},{RET_MS50})", NUM6)

met.cell(row=mrow, column=2, value="Cashflow and cost accounting").font = BOLD
mrow += 1
contrib_ref = m("Total contributed = capital + Σ(cashflow)", f"={param_cells['Initial capital']}+SUM({CASHFLOW_1_4})", MONEY)
withdraw_ref = m("Total withdrawn", "=0", MONEY)
costs_ref = m("Total costs = Σ(rebalance cost)", f"=SUM({COST_1_4})", MONEY)

met.column_dimensions["B"].width = 48
met.column_dimensions["C"].width = 16

# ------------------------------------------------------------ Comparison --
cmp_ws = wb.create_sheet("Comparison")
cmp_ws["B2"] = "Comparison — Excel (independent) vs App (real API response), 6 decimal places"
cmp_ws["B2"].font = TITLE_FONT
cmp_ws["B3"] = f"App output pulled verbatim from docs/verification/4.3b-result.json (run {result['run_id']})."

headers = ["Metric", "Excel value", "App value", "Diff (rounded 6dp)", "Match?"]
hrow = 5
for i, h in enumerate(headers):
    c = cmp_ws.cell(row=hrow, column=2 + i, value=h)
    c.font = BOLD
    c.fill = HEADER_FILL

risk_by_metric = {row["metric"]: row["value"] for row in result["risk_metrics"]["rows"]}
div_corr = result["diversification"]["rows"][0]["correlation"]
asset_by_id = {row["proj_id"]: row for row in result["asset_metrics"]["rows"]}
kset_metrics = asset_by_id["M0209_2548"]
ms50_metrics = asset_by_id["M0155_2547"]

comparisons = [
    ("TWRR", twrr_ref, result["summary"]["twrr"]),
    ("TWRR CAGR", cagr_ref, result["summary"]["twrr_cagr"]),
    ("Volatility", vol_ref, result["summary"]["volatility"]),
    ("Sharpe ratio", sharpe_ref, result["summary"]["sharpe"]),
    ("Sortino ratio", sortino_ref, result["summary"]["sortino"]),
    ("Calmar ratio", calmar_ref, result["summary"]["calmar"]),
    ("Value at Risk 95%", var95_ref, result["summary"]["var_95"]),
    ("Value at Risk 99%", var99_ref, result["summary"]["var_99"]),
    ("Maximum drawdown", mdd_ref, result["summary"]["max_drawdown"]),
    ("IRR (money-weighted)", irr_ref, result["summary"]["irr"]),
    ("Benchmark excess return", excess_ref, result["summary"]["benchmark_excess_return"]),
    ("Beta", beta_ref, risk_by_metric["beta"]),
    ("Alpha", alpha_ref, risk_by_metric["alpha"]),
    ("Tracking error", te_ref, risk_by_metric["tracking_error"]),
    ("Information ratio", ir_ref, risk_by_metric["information_ratio"]),
    ("Correlation (portfolio, benchmark)", corr_bench_ref, risk_by_metric["correlation"]),
    ("Correlation (K-SET50, M-S50)", asset_corr_ref, div_corr),
    ("K-SET50 CAGR", kset_cagr_ref, kset_metrics["cagr"]),
    ("K-SET50 volatility", kset_vol_ref, kset_metrics["volatility"]),
    ("M-S50 CAGR", ms50_cagr_ref, ms50_metrics["cagr"]),
    ("M-S50 volatility", ms50_vol_ref, ms50_metrics["volatility"]),
    ("Total contributed", contrib_ref, result["summary"]["total_contributed"]),
    ("Total withdrawn", withdraw_ref, result["summary"]["total_withdrawn"]),
    ("Total costs", costs_ref, result["summary"]["total_costs"]),
]

row = hrow + 1
for label, excel_ref, app_value in comparisons:
    cmp_ws.cell(row=row, column=2, value=label)
    excel_cell = cmp_ws.cell(row=row, column=3, value=f"={excel_ref}")
    excel_cell.font = GREEN
    excel_cell.number_format = "0.000000"
    app_cell = cmp_ws.cell(row=row, column=4, value=app_value)
    app_cell.font = BLUE
    app_cell.number_format = "0.000000"
    diff_cell = cmp_ws.cell(row=row, column=5, value=f"=ROUND(C{row}-D{row},6)")
    diff_cell.font = BLACK
    diff_cell.number_format = "0.000000"
    match_cell = cmp_ws.cell(row=row, column=6, value=f'=IF(E{row}=0,"MATCH","MISMATCH")')
    match_cell.font = BLACK
    row += 1

# Ending value and per-period equity curve, monthly returns, cashflows, rebalances
row += 1
cmp_ws.cell(row=row, column=2, value="Equity curve (per-period ending value)").font = BOLD
row += 1
for i, point in enumerate(result["equity_curve"]):
    L = get_column_letter(col_for(i))
    cmp_ws.cell(row=row, column=2, value=point["date"])
    excel_cell = cmp_ws.cell(row=row, column=3, value=f"=Simulation!{L}{rows['EndingValue']}")
    excel_cell.font = GREEN
    excel_cell.number_format = MONEY
    app_cell = cmp_ws.cell(row=row, column=4, value=point["value"])
    app_cell.font = BLUE
    app_cell.number_format = MONEY
    cmp_ws.cell(row=row, column=5, value=f"=ROUND(C{row}-D{row},6)").number_format = "0.000000"
    cmp_ws.cell(row=row, column=6, value=f'=IF(E{row}=0,"MATCH","MISMATCH")')
    row += 1

row += 1
cmp_ws.cell(row=row, column=2, value="Rebalance turnover & cost (per event)").font = BOLD
row += 1
for i, reb in enumerate(result["rebalances"]):
    t = i + 1
    L = get_column_letter(col_for(t))
    # The API's "turnover" field is the *ratio* (fraction of portfolio
    # value), not the dollar money_turnover -- compare against the
    # TurnoverRatio row, not MoneyTurnover.
    cmp_ws.cell(row=row, column=2, value=f"{reb['date']} turnover ratio")
    excel_cell = cmp_ws.cell(row=row, column=3, value=f"=Simulation!{L}{rows['TurnoverRatio']}")
    excel_cell.font = GREEN
    excel_cell.number_format = "0.000000"
    app_cell = cmp_ws.cell(row=row, column=4, value=reb["turnover"])
    app_cell.font = BLUE
    app_cell.number_format = "0.000000"
    cmp_ws.cell(row=row, column=5, value=f"=ROUND(C{row}-D{row},6)").number_format = "0.000000"
    cmp_ws.cell(row=row, column=6, value=f'=IF(E{row}=0,"MATCH","MISMATCH")')
    row += 1
    cmp_ws.cell(row=row, column=2, value=f"{reb['date']} cost")
    excel_cell = cmp_ws.cell(row=row, column=3, value=f"=Simulation!{L}{rows['Cost']}")
    excel_cell.font = GREEN
    excel_cell.number_format = MONEY
    app_cell = cmp_ws.cell(row=row, column=4, value=reb["cost"])
    app_cell.font = BLUE
    app_cell.number_format = MONEY
    cmp_ws.cell(row=row, column=5, value=f"=ROUND(C{row}-D{row},6)").number_format = "0.000000"
    cmp_ws.cell(row=row, column=6, value=f'=IF(E{row}=0,"MATCH","MISMATCH")')
    row += 1

last_data_row = row - 1
cmp_ws.column_dimensions["B"].width = 40
for col in "CDEF":
    cmp_ws.column_dimensions[col].width = 16

# ----------------------------------------------------------------- Checks -
chk = wb.create_sheet("Checks")
chk["B2"] = "Checks"
chk["B2"].font = TITLE_FONT
chk["B4"] = "Matches"
chk["C4"] = f'=COUNTIF(Comparison!F{hrow+1}:F{last_data_row},"MATCH")'
chk["B5"] = "Mismatches"
chk["C5"] = f'=COUNTIF(Comparison!F{hrow+1}:F{last_data_row},"MISMATCH")'
chk["B6"] = "Total comparisons"
chk["C6"] = "=C4+C5"
chk["B7"] = "All match?"
chk["C7"] = '=IF(C5=0,"TRUE","FALSE")'
for r in range(4, 8):
    chk.cell(row=r, column=3).font = BLACK
chk.column_dimensions["B"].width = 24

wb.save(OUT_PATH)
print("Workbook built:", OUT_PATH)
print("Comparison rows:", hrow + 1, "to", last_data_row)

